# -*- coding: utf-8 -*-
import os
import re
import logging
import openpyxl
import xlrd
import json
import datetime
class ExcelHandler():
	def __init__(self,folder_path=None):
		self.folder = folder_path
		# pass

	def get_filenames(self,folder_path = '',incld_path = True,re_filter = None):
		'''
		incld_path: 输出是否包含目录
		re_filter : 文件名过滤器,用正则表达式
		# ext_list: 允许的后缀名,不用带"点""
		'''
		if folder_path == '':
			folder_path = self.folder
		else:
			self.folder = folder_path

		for root,dirs,files in os.walk(folder_path):
			results = []
			if re_filter:
				if incld_path:
					results = [os.path.join(root,f) for f in files if re.search(re_filter,f) ]
				else:
					results = [f for f in files if re.search(re_filter,f)]
			else:
				if incld_path:
					results = [os.path.join(root,f) for f in files]
				else:
					results = files
			self.files = results
			return results

	def get_data_from_xls(self,sheet_filter = None,cells = []):
		'''
		从xls表格中读取数据
		sheet_filter 是sheet名过滤器
		cell 是单元格描述数组
		cell[i] = (name,row,col)
		'''
		data_wbs = []
		for file in self.files:
			if re.search(r'.xls$',file):
				try:
					wb = xlrd.open_workbook(file)
				except:
					logging.error('fail to open ',file)
					continue
				data_wb = DataWorkbook(file)
				sht_list = []
				for sht in wb.sheets():
					if sheet_filter:
						if not re.search(sheet_filter,sht.name):
							continue
					data_sheet = DataSheet(sht.name)
					cell_list = []
					for name,row,col in cells: 
						try:
							cell_type = sht.cell_type(row,col)
							if cell_type == 5:#公式错误
								cell_value = "Error"
							else:
								cell_value = sht.cell_value(row,col)
							
						except:
							logging.error('array index out of range for cell(%s,%d,%d) in sheet %s in file ' %(name,row,col,sht.name))
							cell_value = None
						cell_list.append(DataCell(name,row,col,cell_value))
					data_sheet.cells = cell_list
					sht_list.append(data_sheet)
				data_wb.sheets = sht_list
				data_wbs.append(data_wb)
		self.data_wbs = data_wbs
		return data_wb

	def save_to_json(self,file_name="",obj = None):
		"""
		将对象(instance)数据存成json格式
		file_name 保存的文件名,如果省略,则输出按时间命名的json文件
		obj 要被转换的对象,如果省略,则转换遍历文件抓取的工作簿数据
		"""
		if file_name == "":
			file_name = "JSON" + self.get_datetime_str() + ".json"
		if obj == None:
			obj = self.data_wbs
		json_str = json.dumps(obj,default = lambda o:o.__dict__,indent = 4,ensure_ascii = False)
		self.save_to_file(result_file_name = file_name,content = json_str)
	
	def save_to_file(self,result_file_name = "",content ="",sub_folder_name = "outputs"):
		"""
		将content内容(是一个字符串),存入到命名的txt文件中
		如果省略名称,则用时间自动命名
		"""
		cur_dir = os.path.dirname(os.path.abspath(__file__))
		folder_to_save = os.path.join(cur_dir,sub_folder_name)
		if not os.path.exists(folder_to_save):
			os.mkdir(folder_to_save)

		if result_file_name == "":
			result_file_name = 'Data' + self.get_datetime_str() +".txt"
		with open(os.path.join(folder_to_save,result_file_name),"w") as result_file:
			result_file.write(content.encode('utf-8'))

		print "Saved to file: " + result_file_name
		return result_file_name
	
	def get_datetime_str(self):
		return datetime.datetime.now().strftime('%Y%m%d_%H-%M-%S-%f')

	def print_data(self,to_print = True,to_save = True,separator ="|"):
		content = ""
		title_line =  u'file_name'+ separator +separator.join(["%s" %cell.name for cell in self.data_wbs[0].sheets[0].cells])
		# print title_line
		content += title_line+"\n"
		for wb in self.data_wbs:
			for sht in wb.sheets:
				data_line = wb.name + separator +separator.join("%s" %cell.value for cell in sht.cells)
				data_line = data_line.replace('\n','')
				content += data_line+'\n'
		if to_print:
			print content
		if to_save:
			self.save_to_file(content = content)		
		return content

	def write_data_to_xlsx(self,folder = "",wb_filter = r".xlsx$",sheet_filter = None,cells=[]):
		"""
		向指定目录下的xlsx文件中写入cells中描述的数据
		folder 指定目录,如果不填,就用初始化对象的folder
		wb_filter,sheet_filter是工作簿和工作表名称的过滤器,需要填写正则表达式
		wb_filter默认筛出xlsx文件
		cells要写入的数据,是一个数组,数组元素格式为(value="",row,col)
		cells中的行和列是从1开始的,因为是xlsx文件,用openpyxl处理的

		"""
		if folder == "":
			folder = self.folder
		else:
			self.folder = folder

		file_names = self.get_filenames(folder_path = folder,incld_path = True,re_filter = wb_filter)
		for file_name in file_names:
			logging.warning('opening '+file_name)
			wb = openpyxl.load_workbook(file_name)
			for sht in wb.worksheets:
				if sheet_filter:
					if not re.search(sheet_filter,sht.title):
						continue
				logging.warning('working on '+sht.title)
				for value,row,col in cells:
					sht.cell(row = row,column = col).value = value

			wb.save(file_name)
			logging.error(file_name+" saved!")

	def get_and_print_data(self,folder,file_filter = r".xls$", sheet_filter = None, cells=[],to_print = False,to_save = True,to_json = True):
		'自动完成从表格中抓取数据的一系列操作,建立了一个instance之后,可以直接调用'
		# handler = ExcelHandler()
		self.get_filenames(folder_path = folder,re_filter = file_filter)
		self.get_data_from_xls(sheet_filter = sheet_filter,cells = cells)
		self.print_data(to_print = to_print,to_save = to_save)
		if to_json:
			self.save_to_json()
	def merge_sheets(
		self,folder = "",
		wb_filter = r".xlsx$",
		sheet_filter = None,
		header_rows = 0,
		key_column = 1,
		save_dir = os.path.dirname(os.path.abspath(__file__))):
		"""
		合并Worksheets到新的表格
		header_rows标题行占几行?复制时候可以滤除
		key_column关键列,从表中复制的最大行数,由关键列确定,从1行开始,一直找到关键列不为空的行数.

		"""
		if folder == "":
			folder = self.folder
		else:
			self.folder = folder

		file_names = self.get_filenames(folder_path = folder,incld_path = True,re_filter = wb_filter)
		merged_wb = openpyxl.Workbook()
		merged_sht = merged_wb.active
		#-----------------------------
		#从第一个sheet中粘贴标题行
		# wb_first = openpyxl.load_workbook(file_names[0])
		# for sht_first in wb.worksheets:
		# 	if sheet_filter:
		# 		if not re.search(sheet_filter,sht.title):
		# 			continue

		#-----------------------------
		row_cnt = header_rows + 1
		cnt_sheet = 0
		for file_name in file_names:
			# logging.warning('opening '+file_name)
			wb = openpyxl.load_workbook(file_name)
			for sht in wb.worksheets:
				if sheet_filter:
					if not re.search(sheet_filter,sht.title):
						continue
				(my_dir,wb_name) = os.path.split(file_name)
				ws_name = sht.title
				max_row = self.get_max_row(sht,1)
				max_col = sht.max_column
				
				#一次性添加标题行
				if(cnt_sheet == 0):
					self.copy_range_by_num(
						from_sht = sht,
						to_sht = merged_sht,
						from_first_row = 1,
						from_last_row = header_rows,
						from_first_col = 1,
						from_last_col = max_col,
						to_first_row = 1,
						to_first_col = 1)
				#--------------------

				#开始合并
				self.copy_range_by_num(
					from_sht = sht,
					to_sht = merged_sht,
					from_first_row = header_rows+1,
					from_last_row = max_row,
					from_first_col = 1,
					from_last_col = max_col,
					to_first_row = row_cnt,
					to_first_col = 1)
				row_cnt += max_row - (header_rows + 1) + 14
				print wb_name,
				print u" 的工作表: %s 中,复制了%d列,%d行" %(ws_name,sht.max_column,self.get_max_row(sht,1))
				cnt_sheet += 1
		self.save_as_xlsx(merged_wb,name = "merged_order"+self.get_datetime_str(),save_dir = save_dir)
		



	def save_as_xlsx(self,workbook,name="output",save_dir=""):
		if save_dir == "":
			save_dir = self.get_output_dir()
		new_filename = save_dir+"\\"+ name +'.xlsx'
		print "saved to ",new_filename
		workbook.save(new_filename)

	def get_output_dir(self):
		cur_dir = os.path.dirname(os.path.abspath(__file__))
		folder = "output_" + datetime.datetime.now().strftime('%Y%m%d_%H-%M-%S-%f')
		save_dir = os.path.join(cur_dir,folder)
		if not os.path.exists(save_dir):
			os.mkdir(save_dir)
		return save_dir

	def get_max_row(self,sht,key_column):
		"""
		找到最大行数,依据是key_column这一列不为空
		key_column是从1开始的数字
		"""
		row = 1
		while sht.cell(row = row,column = key_column).value:
			row += 1
		return row - 1

	def copy_range(self,from_sht,to_sht,from_range_str,to_cell_str):
		"""
		复制区域函数
		从from_sht中,将form_range_str区域的内容,复制到to_sheet中,粘贴起点是to_cell_str
		"""
		(first_row,first_col) = self.get_cell_pos(to_cell_str)
		row_cnt = first_row
		col_cnt = first_col


		for row in from_sht.iter_rows(from_range_str):
			for cell in row:
				to_sht.cell(row = row_cnt,column = col_cnt).value = cell.value
				col_cnt += 1
			row_cnt += 1 
			col_cnt  = first_col

	def copy_range_by_num(self,from_sht,to_sht,
		from_first_row,from_last_row,from_first_col,from_last_col,
		to_first_row,to_first_col):
		"""
		拷贝区域,所有都是用数字表示,行号和列号都是从1开始

		"""
		num_rows = from_last_row - from_first_row + 1
		num_cols = from_last_col - from_first_col + 1
		for row in range(num_rows):
			for col in range(num_cols):
				to_sht.cell(row = to_first_row + row, column = to_first_col + col).value =\
				from_sht.cell(row = from_first_row + row, column = from_first_col + col).value
				to_sht.cell(row = to_first_row + row, column = to_first_col + col).style =\
				from_sht.cell(row = from_first_row + row, column = from_first_col + col).style

		# for row in range(from_first_row,from_last_row+1):
		# 	for col in range(from_first_col,from_last_col+1):
		# 		to_sht.cell(row = to_first_row +).value = from_sht(row = row,column = col).value()

class DataCell():
	def __init__(self,name,row = -1,col = -1,value = None):
		"""
		self定义,注意xlrd中行和列的index是从0开始的,而openpyxl是从1开始的
		"""
		self.name = name
		self.row = row
		self.col = col
		self.value = value
	def __repr__(self):
		return 'class self:name = %s,row = %d,col = %d,value = %s' %(self.name,self.row,self.col,self.value)
	# def
class DataSheet():
	def __init__(self,name,cells = []):
		self.name = name
		self.cells = cells
class DataWorkbook():
	def __init__(self,name,sheets = []):
		self.name = name.decode('gbk')#处理中文文件名问题
		# self.name = name
		self.sheets = sheets


# def unit_test():
# 	hand = ExcelHandler()
# 	print hand.get_datetime_str()
# 	hand.save_to_file()
# 	hand.save_to_file("sym.txt",u"大哥回答过")
# def unit_test2():
# 	hand = ExcelHandler()
# 	folder = r"E:\kuaipan\github\excel_handler\test_xls"
# 	files = hand.get_filenames(folder,re_filter = r'.xls$')
# 	# for f in files:
# 	# 	print f
# 	hand.get_data_from_xls(
# 		sheet_filter = u'包',
# 		cells = [
# 		(u"num",0,19),
# 		(u"nkt_gm3",1,19),
# 		(u"num_company",2,19),
# 		(u"winner",3,19),
# 		(u"min",4,19),
# 		(u"max",5,19),
# 		(u"average",6,19),
# 		(u"average_no_peak",7,19),
# 		(u"median",8,19),
# 		(u"winner_price",9,19),
# 		(u"nkt_price",10,19)
# 		])

# 	hand.save_to_json()
def get_sheet_names(folder):
	hand = ExcelHandler()
	hand.get_filenames(folder)
	hand.get_data_from_xls()
	for wb in hand.data_wbs:
		for sht in wb.sheets:
			print wb.name,sht.name
def get_bid_data():
	folder = r"C:\Users\adam\Desktop\bidding_xls"
	handler = ExcelHandler()
	handler.get_filenames(folder)

	handler.get_data_from_xls(
		sheet_filter = u"基本信息",
		cells =[
		(u'工作号',1,1),
		(u'开标时间',2,1),
		(u'销售员',3,1),
		(u'项目单位',4,1),
		(u'项目名称',5,1),
		(u'csc专员',6,1),
		# (3.14,6,1),
		])
	handler.save_to_json(u'基本信息'+handler.get_datetime_str()+'.json')
	handler.print_data()

	handler.get_data_from_xls(
		sheet_filter = ur'[(包)(标段)(开标)(门)]',
		cells = [
		(u"num",0,19),
		(u"nkt_gm3",1,19),
		(u"num_company",2,19),
		(u"winner",3,19),
		(u"min",4,19),
		(u"max",5,19),
		(u"average",6,19),
		(u"average_no_peak",7,19),
		(u"median",8,19),
		(u"winner_price",9,19),
		(u"nkt_price",10,19)
		])
	handler.save_to_json(u'包数据'+handler.get_datetime_str()+'.json')
	handler.print_data()
# def test_new_line():

# 	folder = r"C:\Users\adam\Desktop\bidding_xls"
# 	handler = ExcelHandler()
# 	handler.get_filenames(folder,re_filter = r'-010|-011')

# 	handler.get_data_from_xls(
# 		sheet_filter = u"基本信息",
# 		cells =[
# 		(u'工作号',1,1),
# 		(u'开标时间',2,1),
# 		(u'销售员',3,1),
# 		(u'项目单位',4,1),
# 		(u'项目名称',5,1),
# 		(3.14,6,1),
# 		])
# 	# handler.save_to_json(u'基本信息.json')
# 	handler.print_data()
# def unit_test3():
# 	handler = ExcelHandler()
# 	folder = r"E:\kuaipan\github\excel_handler\test_xlsx"
# 	handler.write_data_to_xlsx(folder = folder,
# 		wb_filter = r".xlsx$",
# 		sheet_filter = ur"[(包)(标段)]",
# 		cells=[
# 		(1,22,u"测试数据"),
# 		(2,22,3.14159),
# 		(3,22,"=today()"),
# 		(4,22,"=max(B:B)"),
# 		])
# def unit_test4():
# 	handler = ExcelHandler()
# 	handler.merge_sheets(
# 		folder = r"C:\Users\adam\Desktop\order",
# 		wb_filter = r".xlsx$",
# 		sheet_filter = r"2016",
# 		key_column =1,
# 		header_rows = 1)

if __name__ == "__main__":
	# unit_test4()
	pass
	# get_sheet_names(r"C:\Users\adam\Desktop\bidding_xls")
	# get_bid_data()
	# test_new_line()

